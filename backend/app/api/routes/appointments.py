"""Rutas de citas y pacientes."""
import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session as DBSession

from app.api.deps import get_current_agent
from app.db.models.agent import Agent
from app.db.models.appointment import Appointment
from app.db.models.patient import Patient
from app.db.session import get_db

appointments_router = APIRouter(prefix="/appointments", tags=["appointments"])
patients_router     = APIRouter(prefix="/patients",     tags=["patients"])


# ══════════════════════════════════════════════════════════════════════════════
# CITAS
# ══════════════════════════════════════════════════════════════════════════════

@appointments_router.get("/")
def list_appointments(
    status: str | None = None,
    db: DBSession = Depends(get_db),
    agent: Agent = Depends(get_current_agent),
):
    q = db.query(Appointment).order_by(Appointment.created_at.desc())
    if status:
        q = q.filter(Appointment.status == status)
    items = q.limit(100).all()
    return [
        {
            "id": a.id, "full_name": a.full_name, "phone": a.phone,
            "service": a.service,
            "preferred_date": str(a.preferred_date) if a.preferred_date else None,
            "clinic": a.clinic, "status": a.status,
            "created_at": str(a.created_at),
        }
        for a in items
    ]


@appointments_router.patch("/{appt_id}/confirm")
def confirm_appointment(
    appt_id: int,
    db: DBSession = Depends(get_db),
    agent: Agent = Depends(get_current_agent),
):
    appt = db.query(Appointment).filter(Appointment.id == appt_id).first()
    if not appt:
        raise HTTPException(status_code=404, detail="Cita no encontrada")
    appt.status = "confirmed"
    appt.confirmed_by_agent_id = agent.id
    db.commit()
    return {"ok": True}


@appointments_router.get("/export/excel")
def export_appointments_excel(
    since:  Optional[str] = Query(None, description="Fecha inicio YYYY-MM-DD"),
    until:  Optional[str] = Query(None, description="Fecha fin YYYY-MM-DD"),
    status: Optional[str] = Query(None, description="pending_confirm | confirmed | completed | cancelled"),
    clinic: Optional[str] = Query(None, description="arecibo | bayamon | latam | virtual"),
    db: DBSession = Depends(get_db),
    _: Agent = Depends(get_current_agent),
):
    """
    Exporta citas a Excel en formato listo para carga manual en Vagaro.
    Columnas: Nombre · Teléfono · Servicio · Sede · Fecha · Hora · Estado · Condiciones · Notas
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    since_date = date.fromisoformat(since) if since else None
    until_date = date.fromisoformat(until) if until else None

    q = db.query(Appointment).order_by(
        Appointment.preferred_date.asc(),
        Appointment.preferred_time.asc(),
    )
    if since_date: q = q.filter(Appointment.preferred_date >= since_date)
    if until_date: q = q.filter(Appointment.preferred_date <= until_date)
    if status:     q = q.filter(Appointment.status == status)
    if clinic:     q = q.filter(Appointment.clinic == clinic)

    appointments = q.all()

    # ── Estilos ───────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Citas para Vagaro"

    hf  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    nf  = Font(name="Calibri", size=10)
    bf  = Font(name="Calibri", bold=True, size=10)
    hfl = PatternFill("solid", fgColor="0B4C45")
    gfl = PatternFill("solid", fgColor="F5F1EB")
    brd = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
    )
    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Cabeceras ─────────────────────────────────────────────────────────────
    COLS = [
        ("Nombre completo",     22),
        ("Teléfono",            15),
        ("Servicio",            32),
        ("Sede / Clínica",      18),
        ("Fecha preferida",     16),
        ("Hora preferida",      14),
        ("Estado",              18),
        ("Condiciones médicas", 30),
        ("Notas del agente",    35),
        ("Fecha solicitud",     16),
    ]
    for ci, (header, width) in enumerate(COLS, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = width
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = hf; cell.fill = hfl; cell.border = brd; cell.alignment = ca
    ws.row_dimensions[1].height = 24

    STATUS_ES = {
        "pending_confirm": "Pendiente confirmar",
        "confirmed":       "Confirmada ✅",
        "cancelled":       "Cancelada ❌",
        "completed":       "Completada",
    }
    CLINIC_ES = {
        "arecibo": "Arecibo",
        "bayamon": "Bayamón",
        "latam":   "LATAM / Virtual",
        "virtual": "Virtual",
    }

    # ── Filas ─────────────────────────────────────────────────────────────────
    for ri, a in enumerate(appointments, 2):
        alt = ri % 2 == 0
        row_data = [
            a.full_name or "",
            a.phone or "",
            a.service or "",
            CLINIC_ES.get(a.clinic, a.clinic or ""),
            str(a.preferred_date) if a.preferred_date else "",
            str(a.preferred_time) if a.preferred_time else "",
            STATUS_ES.get(a.status, a.status or ""),
            a.medical_conditions or "",
            a.notes or "",
            str(a.created_at.date()) if a.created_at else "",
        ]
        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = bf if ci == 1 else nf
            cell.border    = brd
            cell.alignment = la
            if alt: cell.fill = gfl

    # ── Fila de totales ───────────────────────────────────────────────────────
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(summary_row, 1, f"Total: {len(appointments)} citas").font = bf
    ws.cell(summary_row, 2, f"Exportado: {date.today()}").font = Font(name="Calibri", size=9, italic=True, color="888888")

    # ── Exportar ──────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output); output.seek(0)

    clinica_txt = f"_{clinic}" if clinic else ""
    filename    = f"Citas_Vagaro{clinica_txt}_{date.today()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# PACIENTES
# ══════════════════════════════════════════════════════════════════════════════

@patients_router.get("/")
def list_patients(
    db: DBSession = Depends(get_db),
    _: Agent = Depends(get_current_agent),
):
    patients = db.query(Patient).order_by(Patient.created_at.desc()).limit(200).all()
    return [
        {
            "id": p.id, "whatsapp_number": p.whatsapp_number, "full_name": p.full_name,
            "location_type": p.location_type, "is_recurrent": bool(p.is_recurrent),
            "last_interaction_at": str(p.last_interaction_at) if p.last_interaction_at else None,
        }
        for p in patients
    ]


@patients_router.get("/{patient_id}")
def get_patient(
    patient_id: int,
    db: DBSession = Depends(get_db),
    _: Agent = Depends(get_current_agent),
):
    p = db.query(Patient).filter(Patient.id == patient_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    appts = (
        db.query(Appointment)
        .filter(Appointment.patient_id == patient_id)
        .order_by(Appointment.created_at.desc())
        .all()
    )
    return {
        "id": p.id, "whatsapp_number": p.whatsapp_number, "full_name": p.full_name,
        "birth_date":       str(p.birth_date) if p.birth_date else None,
        "email":            p.email,
        "location_type":    p.location_type,
        "is_recurrent":     bool(p.is_recurrent),
        "notes":            p.notes,
        "appointments": [
            {"id": a.id, "service": a.service, "status": a.status, "created_at": str(a.created_at)}
            for a in appts
        ],
    }