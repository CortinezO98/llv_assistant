"""
app/api/routes/reports.py

Reportería mensual para el equipo de soporte LRV.
8 KPIs requeridos + exportación Excel y PDF.
"""
import io
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import cast, func, String
from sqlalchemy.orm import Session as DBSession

from app.api.deps import get_current_agent
from app.db.models.agent import Agent
from app.db.models.appointment import Appointment
from app.db.models.analytics import AnalyticsEvent
from app.db.models.patient import Patient
from app.db.models.payment import Payment
from app.db.models.session import Session
from app.db.session import get_db
from app.schemas.reports import ReportFilters
from app.services.report_service import build_report_data, parse_products_csv

router = APIRouter(prefix="/reports", tags=["reports"])




def _parse_products(products: str | None) -> list[str]:
    return parse_products_csv(products)


def _build_report_data(
    db: DBSession,
    since: date,
    until: date,
    channel: str | None = None,
    session_status: str | None = None,
    agent_id: int | None = None,
    products: list[str] | None = None,
    location: str | None = None,
    payment_status: str | None = None,
) -> dict:
    filters = ReportFilters(
        date_from=since,
        date_to=until,
        channel=channel,
        session_status=session_status,
        agent_id=agent_id,
        products=products or [],
        location=location,
        payment_status=payment_status,
    )
    return build_report_data(db, filters)

@router.get("/summary")
def get_report_summary(
    days: int = 30,
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    channel: str | None = Query(default=None),
    session_status: str | None = Query(default=None),
    agent_id: int | None = Query(default=None),
    products: str | None = Query(default=None),
    location: str | None = Query(default=None),
    payment_status: str | None = Query(default=None),
    db: DBSession = Depends(get_db),
    _: Agent = Depends(get_current_agent),
):
    until = date_to or date.today()
    since = date_from or (until - timedelta(days=days))
    return _build_report_data(db, since, until, channel, session_status, agent_id, _parse_products(products), location, payment_status)


@router.get("/export/excel")
def export_excel(
    days: int = 30,
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    channel: str | None = Query(default=None),
    session_status: str | None = Query(default=None),
    agent_id: int | None = Query(default=None),
    products: str | None = Query(default=None),
    location: str | None = Query(default=None),
    payment_status: str | None = Query(default=None),
    db: DBSession = Depends(get_db),
    _: Agent = Depends(get_current_agent),
):
    """Exporta el reporte mensual en formato Excel."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    until = date_to or date.today()
    since = date_from or (until - timedelta(days=days))
    data = _build_report_data(db, since, until, channel, session_status, agent_id, _parse_products(products), location, payment_status)

    wb = openpyxl.Workbook()

    # ── Colores ───────────────────────────────────────────────────────────────
    GREEN_DARK  = "1B5E3F"
    GREEN_MED   = "2E7D52"
    GREEN_LIGHT = "EAF4EF"
    WHITE       = "FFFFFF"
    GRAY        = "F5F5F5"

    header_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    title_font   = Font(name="Calibri", bold=True, color=GREEN_DARK, size=14)
    normal_font  = Font(name="Calibri", size=10)
    bold_font    = Font(name="Calibri", bold=True, size=10)
    header_fill  = PatternFill("solid", fgColor=GREEN_DARK)
    subhdr_fill  = PatternFill("solid", fgColor=GREEN_MED)
    light_fill   = PatternFill("solid", fgColor=GREEN_LIGHT)
    gray_fill    = PatternFill("solid", fgColor=GRAY)
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:   cell.font = font
        if fill:   cell.fill = fill
        if align:  cell.alignment = align
        if border: cell.border = border
        return cell

    def section_header(ws, row, col, text, width=2):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
        set_cell(ws, row, col, text, font=Font(name="Calibri", bold=True, color=WHITE, size=11), fill=subhdr_fill, align=center)

    def kv_row(ws, row, label, value, alt=False):
        set_cell(ws, row, 1, label, font=bold_font, fill=gray_fill if alt else PatternFill(), align=left, border=thin_border)
        set_cell(ws, row, 2, value, font=normal_font, fill=gray_fill if alt else PatternFill(), align=left, border=thin_border)

    # ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 22

    # Título
    ws1.merge_cells("A1:B1")
    set_cell(ws1, 1, 1, "LLV ASSISTANT — REPORTE MENSUAL", font=title_font, align=center)
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:B2")
    set_cell(ws1, 2, 1, f"Período: {data['period']['since']}  →  {data['period']['until']}", font=Font(name="Calibri", size=10, color="666666"), align=center)
    ws1.merge_cells("A3:B3")
    set_cell(ws1, 3, 1, "LRV Aesthetic & Wellness Clinic", font=Font(name="Calibri", size=10, italic=True, color="888888"), align=center)

    row = 5
    # 1. Conversaciones
    section_header(ws1, row, 1, "1. CONVERSACIONES E USUARIOS", 2); row += 1
    kv_row(ws1, row, "Total conversaciones iniciadas", data["conversations"]["total"]); row += 1
    kv_row(ws1, row, "Usuarios únicos", data["conversations"]["unique_users"], True); row += 1
    kv_row(ws1, row, "Conversaciones completadas", data["conversations"]["completed"]); row += 1
    kv_row(ws1, row, "% Completadas", f"{data['conversations']['pct_completed']}%", True); row += 1

    row += 1
    # 2. Abandono
    section_header(ws1, row, 1, "2. PUNTOS DE ABANDONO", 2); row += 1
    status_labels = {"active": "Activas", "in_agent": "Con agente", "completed": "Completadas", "closed": "Cerradas"}
    for status, count in data["conversations"]["status_distribution"].items():
        alt = (row % 2 == 0)
        kv_row(ws1, row, status_labels.get(status, status), count, alt); row += 1

    row += 1
    # 3. Agentes
    section_header(ws1, row, 1, "3. PASO A ASESOR HUMANO", 2); row += 1
    kv_row(ws1, row, "Total escaladas a agente", data["agents"]["escalated"]); row += 1
    kv_row(ws1, row, "% del total de conversaciones", f"{data['agents']['pct_escalated']}%", True); row += 1
    for agent_data in data["agents"]["by_agent"]:
        kv_row(ws1, row, f"  → {agent_data['name']}", agent_data["count"]); row += 1

    row += 1
    # 4. Citas y ventas
    section_header(ws1, row, 1, "4. CONVERSIÓN A CITAS Y VENTAS", 2); row += 1
    kv_row(ws1, row, "Citas solicitadas", data["appointments"]["total_requested"]); row += 1
    kv_row(ws1, row, "Citas confirmadas", data["appointments"]["confirmed"], True); row += 1
    kv_row(ws1, row, "% Conversión a citas", f"{data['appointments']['conversion_pct']}%"); row += 1
    kv_row(ws1, row, "Ventas verificadas (pagos)", data["sales"]["verified_payments"], True); row += 1

    row += 1
    # 5. Ingresos
    section_header(ws1, row, 1, "5. INGRESOS DEL CANAL", 2); row += 1
    kv_row(ws1, row, "Total ingresos verificados (USD)", f"${data['sales']['total_revenue_usd']:.2f}"); row += 1
    for method, count in data["sales"]["payment_methods"].items():
        kv_row(ws1, row, f"  → {method}", f"{count} pagos", True); row += 1

    row += 1
    # 6. Canales
    section_header(ws1, row, 1, "6. CANALES DE ENTRADA", 2); row += 1
    for channel, count in data["channels"].items():
        kv_row(ws1, row, channel.capitalize(), count); row += 1

    row += 1
    # 7. Satisfacción
    section_header(ws1, row, 1, "7. SATISFACCIÓN / FEEDBACK", 2); row += 1
    kv_row(ws1, row, "Puntaje promedio", data["satisfaction"]["note"]); row += 1

    row += 1
    # 8. Pacientes
    section_header(ws1, row, 1, "8. PACIENTES NUEVOS Y RECURRENTES", 2); row += 1
    kv_row(ws1, row, "Pacientes nuevos este período", data["patients"]["new_this_period"]); row += 1
    kv_row(ws1, row, "Total pacientes recurrentes", data["patients"]["total_recurrent"], True); row += 1

    # ── Hoja 2: Servicios más solicitados ─────────────────────────────────────
    ws2 = wb.create_sheet("Top Servicios")
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 15

    ws2.merge_cells("A1:B1")
    set_cell(ws2, 1, 1, "TOP SERVICIOS MÁS SOLICITADOS", font=title_font, fill=light_fill, align=center)

    set_cell(ws2, 2, 1, "Servicio", font=header_font, fill=header_fill, align=center, border=thin_border)
    set_cell(ws2, 2, 2, "Solicitudes", font=header_font, fill=header_fill, align=center, border=thin_border)

    for i, svc in enumerate(data["appointments"]["top_services"], start=3):
        alt = (i % 2 == 0)
        set_cell(ws2, i, 1, svc["service"], font=normal_font, fill=gray_fill if alt else PatternFill(), align=left, border=thin_border)
        set_cell(ws2, i, 2, svc["count"], font=bold_font, fill=gray_fill if alt else PatternFill(), align=center, border=thin_border)

    if not data["appointments"]["top_services"]:
        ws2.merge_cells("A3:B3")
        set_cell(ws2, 3, 1, "Sin datos en este período", font=normal_font, align=center)

    # ── Guardar ───────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"LLV_Reporte_{since}_{until}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pdf")
def export_pdf(
    days: int = 30,
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    channel: str | None = Query(default=None),
    session_status: str | None = Query(default=None),
    agent_id: int | None = Query(default=None),
    products: str | None = Query(default=None),
    location: str | None = Query(default=None),
    payment_status: str | None = Query(default=None),
    db: DBSession = Depends(get_db),
    _: Agent = Depends(get_current_agent),
):
    """Exporta el reporte mensual en formato PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    until = date_to or date.today()
    since = date_from or (until - timedelta(days=days))
    data = _build_report_data(db, since, until, channel, session_status, agent_id, _parse_products(products), location, payment_status)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    GREEN = colors.HexColor("#1B5E3F")
    GREEN_LIGHT = colors.HexColor("#EAF4EF")
    GREEN_MED = colors.HexColor("#2E7D52")
    GRAY = colors.HexColor("#F5F5F5")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], textColor=GREEN, fontSize=18, spaceAfter=4)
    sub_style   = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.HexColor("#666666"), fontSize=10, spaceAfter=12)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], textColor=GREEN_MED, fontSize=12, spaceBefore=16, spaceAfter=6)

    def make_table(header, rows):
        col_widths = [3.8*inch, 2.5*inch]
        table_data = [[Paragraph(f"<b>{h}</b>", styles["Normal"]) for h in header]] + [
            [Paragraph(str(r[0]), styles["Normal"]), Paragraph(str(r[1]), styles["Normal"])]
            for r in rows
        ]
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRAY]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    elements = []

    # Encabezado
    elements.append(Paragraph("LLV ASSISTANT — Reporte Mensual", title_style))
    elements.append(Paragraph(f"Período: {data['period']['since']} → {data['period']['until']}  |  LRV Aesthetic & Wellness Clinic", sub_style))

    # 1. Conversaciones
    elements.append(Paragraph("1. Conversaciones e Usuarios", section_style))
    elements.append(make_table(["Métrica", "Valor"], [
        ["Total conversaciones iniciadas", data["conversations"]["total"]],
        ["Usuarios únicos", data["conversations"]["unique_users"]],
        ["Conversaciones completadas", data["conversations"]["completed"]],
        ["% Completadas", f"{data['conversations']['pct_completed']}%"],
    ]))

    # 2. Abandono
    elements.append(Paragraph("2. Puntos de Abandono", section_style))
    status_labels = {"active": "Activas", "in_agent": "Con agente", "completed": "Completadas", "closed": "Cerradas"}
    abandonment_rows = [[status_labels.get(s, s), c] for s, c in data["conversations"]["status_distribution"].items()]
    elements.append(make_table(["Estado", "Cantidad"], abandonment_rows or [["Sin datos", "—"]]))

    # 3. Agentes
    elements.append(Paragraph("3. Paso a Asesor Humano", section_style))
    agent_rows = [
        ["Total escaladas a agente", data["agents"]["escalated"]],
        ["% del total", f"{data['agents']['pct_escalated']}%"],
    ] + [[f"→ {a['name']}", a["count"]] for a in data["agents"]["by_agent"]]
    elements.append(make_table(["Métrica / Agente", "Valor"], agent_rows))

    # 4. Citas y ventas
    elements.append(Paragraph("4. Conversión a Citas y Ventas", section_style))
    elements.append(make_table(["Métrica", "Valor"], [
        ["Citas solicitadas", data["appointments"]["total_requested"]],
        ["Citas confirmadas", data["appointments"]["confirmed"]],
        ["% Conversión a citas", f"{data['appointments']['conversion_pct']}%"],
        ["Ventas verificadas (pagos)", data["sales"]["verified_payments"]],
    ]))

    # 5. Ingresos
    elements.append(Paragraph("5. Ingresos del Canal", section_style))
    method_rows = [[m, f"{c} pagos"] for m, c in data["sales"]["payment_methods"].items()]
    elements.append(make_table(["Métrica", "Valor"], [
        ["Total ingresos verificados (USD)", f"${data['sales']['total_revenue_usd']:.2f}"],
    ] + method_rows))

    # 6. Canales
    elements.append(Paragraph("6. Canales de Entrada", section_style))
    elements.append(make_table(["Canal", "Conversaciones"],
        [[c.capitalize(), n] for c, n in data["channels"].items()] or [["WhatsApp", data["conversations"]["total"]]]))

    # 7. Satisfacción
    elements.append(Paragraph("7. Satisfacción / Feedback", section_style))
    elements.append(make_table(["Métrica", "Estado"], [
        ["Encuesta post-conversación", "Pendiente de implementar"],
        ["Puntaje promedio", "—"],
    ]))

    # 8. Pacientes
    elements.append(Paragraph("8. Pacientes Nuevos y Recurrentes", section_style))
    elements.append(make_table(["Métrica", "Valor"], [
        ["Pacientes nuevos este período", data["patients"]["new_this_period"]],
        ["Total pacientes recurrentes", data["patients"]["total_recurrent"]],
    ]))

    # Top servicios
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph("Top Servicios más Solicitados", section_style))
    svc_rows = [[s["service"], s["count"]] for s in data["appointments"]["top_services"]]
    elements.append(make_table(["Servicio", "Solicitudes"], svc_rows or [["Sin datos", "—"]]))

    doc.build(elements)
    output.seek(0)

    filename = f"LLV_Reporte_{since}_{until}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
